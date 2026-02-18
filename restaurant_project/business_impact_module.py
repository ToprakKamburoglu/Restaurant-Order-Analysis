
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def calculate_business_impact(df, recommendations_df, restaurant_pricing):
    """
    Calculate comprehensive business impact of combo menu implementations

    Parameters:
    - df: Cleaned order DataFrame
    - recommendations_df: Combo recommendations DataFrame
    - restaurant_pricing: Dict with avg order values per restaurant

    Returns:
    - impact_df: DataFrame with business metrics
    """

    impact_results = []

    for _, rec in recommendations_df.iterrows():
        restaurant = rec['Restaurant']
        combo_items = rec['Combo Items']
        lift = rec['Lift']
        confidence = rec['Confidence (%)'] / 100
        support = rec['Support (%)'] / 100
        occurrences = rec['Occurrences']
        discount = rec['Discount Recommendation']

        # Get restaurant data
        rest_df = df[df['Restaurant name'] == restaurant]
        total_orders = len(rest_df)
        avg_order_value = restaurant_pricing[restaurant]['avg_total']
        avg_subtotal = restaurant_pricing[restaurant]['avg_subtotal']

        # CALCULATION 1: Potential Additional Sales
        # If we promote this combo, how many more people will buy it?
        current_adoption = occurrences  # Currently buying this combo

        # Potential new customers (conservative: 20% of customers who buy one item will try combo)
        potential_new_customers = int(total_orders * support * 0.20)

        # CALCULATION 2: Revenue Impact
        # Assume combo items together cost more than average order
        combo_value_multiplier = 1.5  # Combo typically 50% higher value
        estimated_combo_value = avg_order_value * combo_value_multiplier

        # Parse discount percentage
        if '-' in discount:
            discount_low = float(discount.split('-')[0].replace('%', '')) / 100
            discount_high = float(discount.split('-')[1].replace('%', '')) / 100
            avg_discount = (discount_low + discount_high) / 2
        else:
            avg_discount = float(discount.replace('%', '')) / 100

        # Revenue per combo after discount
        combo_revenue = estimated_combo_value * (1 - avg_discount)

        # Monthly projections (assuming current is monthly data)
        monthly_additional_revenue = combo_revenue * potential_new_customers
        annual_additional_revenue = monthly_additional_revenue * 12

        # CALCULATION 3: Customer Satisfaction Impact
        # Higher lift = stronger association = higher satisfaction when offered together
        satisfaction_score = min(lift * 25, 100)  # Scale lift to 0-100

        # CALCULATION 4: Operational Efficiency
        # Combo orders are easier to prepare (items often complement each other)
        prep_time_reduction = f"{min(lift * 5, 15):.1f}%"  # Max 15% reduction

        # CALCULATION 5: Customer Retention
        # Customers who buy combos are more likely to return
        retention_boost = f"{min(confidence * 20, 30):.1f}%"  # Max 30% boost

        impact_results.append({
            'Restaurant': restaurant,
            'Combo Items': combo_items,
            'Priority': rec['Priority'],
            'Lift': lift,
            'Current Monthly Sales': occurrences,
            'Potential New Customers/Month': potential_new_customers,
            'Growth Potential (%)': round((potential_new_customers / max(occurrences, 1)) * 100, 1),
            'Estimated Combo Value ($)': round(estimated_combo_value, 2),
            'Discount Offered': discount,
            'Revenue per Combo ($)': round(combo_revenue, 2),
            'Additional Monthly Revenue ($)': round(monthly_additional_revenue, 2),
            'Additional Annual Revenue ($)': round(annual_additional_revenue, 2),
            'Customer Satisfaction Score': round(satisfaction_score, 1),
            'Prep Time Reduction': prep_time_reduction,
            'Customer Retention Boost': retention_boost,
            'ROI Category': 'High' if monthly_additional_revenue > 10000 else 'Medium' if monthly_additional_revenue > 5000 else 'Low'
        })

    return pd.DataFrame(impact_results)


def create_business_summary(impact_df):
    """
    Create executive summary of business impact
    """

    summary_data = []

    for restaurant in impact_df['Restaurant'].unique():
        rest_impact = impact_df[impact_df['Restaurant'] == restaurant]

        total_monthly_revenue = rest_impact['Additional Monthly Revenue ($)'].sum()
        total_annual_revenue = rest_impact['Additional Annual Revenue ($)'].sum()
        avg_satisfaction = rest_impact['Customer Satisfaction Score'].mean()
        high_roi_combos = len(rest_impact[rest_impact['ROI Category'] == 'High'])
        total_combos = len(rest_impact)

        summary_data.append({
            'Restaurant': restaurant,
            'Total Combo Opportunities': total_combos,
            'High ROI Combos': high_roi_combos,
            'Projected Monthly Revenue ($)': round(total_monthly_revenue, 2),
            'Projected Annual Revenue ($)': round(total_annual_revenue, 2),
            'Avg Customer Satisfaction Score': round(avg_satisfaction, 1),
            'Implementation Priority': 'Immediate' if total_monthly_revenue > 20000 else 'High' if total_monthly_revenue > 10000 else 'Medium'
        })

    return pd.DataFrame(summary_data)


def save_business_impact_excel(impact_df, summary_df, filename='06_Business_Impact_Analysis.xlsx'):
    """
    Save business impact analysis to Excel with rich formatting
    """
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook

    wb = Workbook()

    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"

    # Add title
    ws1['A1'] = 'BUSINESS IMPACT ANALYSIS - EXECUTIVE SUMMARY'
    ws1['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws1['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws1.merge_cells('A1:G1')

    ws1['A2'] = f'Generated: {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws1['A2'].font = Font(italic=True)

    # Add summary data
    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 4):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 4:  # Header row
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for col_idx, column in enumerate(ws1.columns, 1):
        max_length = 0
        for cell in column:
            try:
                # Skip merged cells
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        if max_length > 0:
            from openpyxl.utils import get_column_letter
            ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 50)

    # Sheet 2: Detailed Impact
    ws2 = wb.create_sheet("Detailed Impact")

    ws2['A1'] = 'COMBO MENU - DETAILED BUSINESS IMPACT ANALYSIS'
    ws2['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws2['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws2.merge_cells('A1:P1')

    # Add impact data
    for r_idx, row in enumerate(dataframe_to_rows(impact_df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 3:  # Header row
                cell.font = Font(bold=True, color='FFFFFF', size=10)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Color code ROI Category
            if c_idx == impact_df.columns.get_loc('ROI Category') + 1 and r_idx > 3:
                if value == 'High':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif value == 'Medium':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Auto-adjust column widths
    for col_idx, column in enumerate(ws2.columns, 1):
        max_length = 0
        for cell in column:
            try:
                # Skip merged cells
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        if max_length > 0:
            from openpyxl.utils import get_column_letter
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)

    wb.save(filename)
    print(f"\nBusiness impact analysis saved: {filename}")


def plot_business_impact(impact_df, summary_df):
    """
    Create visualizations for business impact
    """
    fig = plt.figure(figsize=(20, 12))
    gs = fig.add_gridspec(2, 3, hspace=0.3, wspace=0.3)

    fig.suptitle('Business Impact Analysis - Combo Menu Implementation',
                 fontsize=18, fontweight='bold', y=0.98)

    # 1. Revenue Potential by Restaurant
    ax1 = fig.add_subplot(gs[0, 0])
    summary_sorted = summary_df.sort_values('Projected Annual Revenue ($)', ascending=False)
    bars = ax1.barh(summary_sorted['Restaurant'], summary_sorted['Projected Annual Revenue ($)'],
                    color=['#27ae60', '#2ecc71', '#f39c12'], edgecolor='black', linewidth=1.5)
    ax1.set_xlabel('Projected Annual Revenue ($)', fontweight='bold', fontsize=12)
    ax1.set_title('A) Annual Revenue Potential by Restaurant', fontweight='bold', fontsize=13, pad=15)
    ax1.grid(axis='x', alpha=0.3)

    for i, bar in enumerate(bars):
        width = bar.get_width()
        ax1.text(width, bar.get_y() + bar.get_height() / 2,
                 f'${width:,.0f}', ha='left', va='center', fontweight='bold', fontsize=10)

    # 2. ROI Distribution
    ax2 = fig.add_subplot(gs[0, 1])
    roi_counts = impact_df['ROI Category'].value_counts()
    colors_roi = {'High': '#27ae60', 'Medium': '#f39c12', 'Low': '#e74c3c'}
    bars = ax2.bar(roi_counts.index, roi_counts.values,
                   color=[colors_roi[cat] for cat in roi_counts.index],
                   edgecolor='black', linewidth=1.5)
    ax2.set_ylabel('Number of Combos', fontweight='bold', fontsize=12)
    ax2.set_title('B) ROI Category Distribution', fontweight='bold', fontsize=13, pad=15)
    ax2.grid(axis='y', alpha=0.3)

    for bar in bars:
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width() / 2., height,
                 f'{int(height)}', ha='center', va='bottom', fontweight='bold', fontsize=11)

    # 3. Customer Satisfaction vs Revenue
    ax3 = fig.add_subplot(gs[0, 2])
    colors_map = {'Aura Pizzas': '#2ecc71', 'Swaad': '#27ae60', 'Dilli Burger Adda': '#f39c12'}

    for restaurant in impact_df['Restaurant'].unique():
        rest_data = impact_df[impact_df['Restaurant'] == restaurant]
        ax3.scatter(rest_data['Customer Satisfaction Score'],
                    rest_data['Additional Monthly Revenue ($)'],
                    s=rest_data['Lift'] * 50,
                    alpha=0.6,
                    color=colors_map.get(restaurant, '#95a5a6'),
                    label=restaurant,
                    edgecolors='black',
                    linewidth=1)

    ax3.set_xlabel('Customer Satisfaction Score', fontweight='bold', fontsize=12)
    ax3.set_ylabel('Monthly Revenue Potential ($)', fontweight='bold', fontsize=12)
    ax3.set_title('C) Satisfaction vs Revenue Potential', fontweight='bold', fontsize=13, pad=15)
    ax3.legend(fontsize=10)
    ax3.grid(alpha=0.3)

    # 4. Top 10 Revenue Opportunities
    ax4 = fig.add_subplot(gs[1, :2])
    top10 = impact_df.nlargest(10, 'Additional Annual Revenue ($)')

    y_pos = np.arange(len(top10))
    bars = ax4.barh(y_pos, top10['Additional Annual Revenue ($)'],
                    color=['#27ae60' if x == 'High' else '#f39c12' if x == 'Medium' else '#e74c3c'
                           for x in top10['ROI Category']],
                    edgecolor='black', linewidth=1.5)

    ax4.set_yticks(y_pos)
    labels = [f"{row['Restaurant'][:15]}\n{row['Combo Items'][:40]}..."
              if len(row['Combo Items']) > 40 else f"{row['Restaurant'][:15]}\n{row['Combo Items']}"
              for _, row in top10.iterrows()]
    ax4.set_yticklabels(labels, fontsize=9)
    ax4.set_xlabel('Annual Revenue Potential ($)', fontweight='bold', fontsize=12)
    ax4.set_title('D) Top 10 Revenue Opportunities', fontweight='bold', fontsize=13, pad=15)
    ax4.grid(axis='x', alpha=0.3)

    for i, (bar, value) in enumerate(zip(bars, top10['Additional Annual Revenue ($)'])):
        ax4.text(value, bar.get_y() + bar.get_height() / 2,
                 f' ${value:,.0f}', ha='left', va='center', fontweight='bold', fontsize=9)

    # 5. Implementation Priority Matrix
    ax5 = fig.add_subplot(gs[1, 2])
    priority_counts = summary_df['Implementation Priority'].value_counts()
    colors_priority = {'Immediate': '#e74c3c', 'High': '#f39c12', 'Medium': '#3498db'}

    wedges, texts, autotexts = ax5.pie(priority_counts.values,
                                       labels=priority_counts.index,
                                       autopct='%1.0f%%',
                                       colors=[colors_priority[p] for p in priority_counts.index],
                                       startangle=90,
                                       textprops={'fontsize': 11, 'weight': 'bold'})

    for autotext in autotexts:
        autotext.set_color('white')

    ax5.set_title('E) Implementation Priority', fontweight='bold', fontsize=13, pad=15)

    plt.savefig('07_Business_Impact_Visualization.png', dpi=300, bbox_inches='tight')
    print("Business impact visualization saved: 07_Business_Impact_Visualization.png")
    plt.close()


def generate_executive_report(summary_df, impact_df):
    """
    Generate text-based executive report
    """
    print("\n" + "=" * 80)
    print("EXECUTIVE SUMMARY - BUSINESS IMPACT ANALYSIS")
    print("=" * 80)

    total_annual_revenue = summary_df['Projected Annual Revenue ($)'].sum()
    total_monthly_revenue = summary_df['Projected Monthly Revenue ($)'].sum()
    total_combos = summary_df['Total Combo Opportunities'].sum()
    high_roi_combos = summary_df['High ROI Combos'].sum()

    print(f"\nOVERALL IMPACT:")
    print(f"  Total Combo Menu Opportunities: {int(total_combos)}")
    print(f"  High ROI Opportunities: {int(high_roi_combos)} ({(high_roi_combos / total_combos) * 100:.1f}%)")
    print(f"  Projected Monthly Revenue Increase: ${total_monthly_revenue:,.2f}")
    print(f"  Projected Annual Revenue Increase: ${total_annual_revenue:,.2f}")

    print(f"\nBY RESTAURANT:")
    for _, row in summary_df.sort_values('Projected Annual Revenue ($)', ascending=False).iterrows():
        print(f"\n  {row['Restaurant']}:")
        print(f"    Combo Opportunities: {int(row['Total Combo Opportunities'])}")
        print(f"    Annual Revenue Potential: ${row['Projected Annual Revenue ($)']:,.2f}")
        print(f"    Customer Satisfaction: {row['Avg Customer Satisfaction Score']:.1f}/100")
        print(f"    Priority: {row['Implementation Priority']}")

    print(f"\nTOP 3 IMMEDIATE ACTIONS:")
    top3 = impact_df.nlargest(3, 'Additional Annual Revenue ($)')
    for i, (_, row) in enumerate(top3.iterrows(), 1):
        print(f"\n  {i}. {row['Restaurant']} - {row['Combo Items']}")
        print(f"     Annual Revenue: ${row['Additional Annual Revenue ($)']:,.2f}")
        print(f"     Discount: {row['Discount Offered']}")
        print(f"     Customer Satisfaction: {row['Customer Satisfaction Score']:.1f}/100")
        print(f"     Action: Implement {row['Priority']} priority combo menu")

    print("\n" + "=" * 80)


# Example usage function
def run_business_impact_analysis(df, recommendations_df):
    """
    Main function to run complete business impact analysis
    """

    # Calculate restaurant pricing
    restaurant_pricing = {}
    for restaurant in df['Restaurant name'].unique():
        rest_df = df[df['Restaurant name'] == restaurant]
        restaurant_pricing[restaurant] = {
            'avg_total': rest_df['Total'].mean(),
            'avg_subtotal': rest_df['Bill subtotal'].mean()
        }

    # Calculate impact
    print("\nCalculating business impact...")
    impact_df = calculate_business_impact(df, recommendations_df, restaurant_pricing)

    # Create summary
    summary_df = create_business_summary(impact_df)

    # Save to Excel
    save_business_impact_excel(impact_df, summary_df)

    # Create visualizations
    plot_business_impact(impact_df, summary_df)

    # Generate executive report
    generate_executive_report(summary_df, impact_df)

    return impact_df, summary_df