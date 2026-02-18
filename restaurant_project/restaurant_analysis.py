
import pandas as pd
import numpy as np
from mlxtend.preprocessing import TransactionEncoder
from mlxtend.frequent_patterns import apriori, fpgrowth, association_rules
import matplotlib.pyplot as plt
import seaborn as sns
import time
from collections import Counter
import warnings
from business_impact_module import run_business_impact_analysis

warnings.filterwarnings('ignore')

# Visual settings
plt.style.use('seaborn-v0_8-whitegrid')
sns.set_palette("Set2")


# ============================================================================
# HELPER FUNCTIONS FOR EXCEL FORMATTING
# ============================================================================

def save_excel_with_formatting(df, filename, sheet_name='Sheet1', add_metadata=None):
    """
    Save DataFrame to Excel with proper column widths and formatting
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    # Save DataFrame
    df.to_excel(filename, index=False, sheet_name=sheet_name)

    # Load workbook for formatting
    wb = load_workbook(filename)
    ws = wb[sheet_name]

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Set width with some padding
        adjusted_width = min(max_length + 3, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Format header row (Standard Blue Style)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add metadata if provided
    if add_metadata:
        ws.insert_rows(1, len(add_metadata) + 1)

        for i, (key, value) in enumerate(add_metadata.items(), 1):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=value)
            ws.cell(row=i, column=1).font = Font(bold=True)

        # Add separator row
        ws.insert_rows(len(add_metadata) + 1, 1)

    wb.save(filename)
    print(f"Excel file saved with formatting: {filename}")


# ============================================================================
# PART 1: DATA LOADING AND CLEANING
# ============================================================================

def load_and_clean_data(filepath):
    """
    Load and clean the original data

    Returns:
    - df_clean: Cleaned DataFrame
    - metadata: Dictionary with cleaning statistics
    """
    print("=" * 80)
    print("DATA LOADING AND CLEANING")
    print("=" * 80)

    # Load original data
    df_original = pd.read_csv(filepath)
    original_count = len(df_original)

    print(f"\nOriginal data loaded successfully")
    print(f"   Total orders: {original_count:,}")
    print(f"   Number of restaurants: {df_original['Restaurant name'].nunique()}")

    # Order Status distribution
    print(f"\nOrder Status Distribution:")
    status_dist = {}
    for status, count in df_original['Order Status'].value_counts().items():
        pct = (count / original_count) * 100
        status_dist[status] = {'count': count, 'percentage': pct}
        print(f"   {status:20s}: {count:6,} ({pct:5.2f}%)")

    # Filter only "Delivered" orders
    df_clean = df_original[df_original['Order Status'] == 'Delivered'].copy()
    removed = original_count - len(df_clean)
    cleaned_count = len(df_clean)

    print(f"\nFiltering: Only 'Delivered' orders retained")
    print(f"   Orders removed: {removed:,} ({(removed / original_count) * 100:.2f}%)")
    print(f"   Orders remaining: {cleaned_count:,} ({(cleaned_count / original_count) * 100:.2f}%)")

    # Check for null values
    null_items = df_clean['Items in order'].isnull().sum()
    if null_items > 0:
        print(f"\nNull 'Items in order': {null_items:,} orders removed")
        df_clean = df_clean[df_clean['Items in order'].notna()]
    else:
        print(f"\nNull value check: All orders have product information")

    final_count = len(df_clean)
    print(f"\nFinal Data: {final_count:,} orders")

    # Create metadata dictionary
    metadata = {
        'Original Orders': original_count,
        'Delivered Orders': cleaned_count,
        'Removed Orders': removed,
        'Removal Rate (%)': round((removed / original_count) * 100, 2),
        'Final Dataset': final_count,
        'Data Retention (%)': round((final_count / original_count) * 100, 2)
    }

    return df_clean, metadata


# ============================================================================
# PART 2: EXPLORATORY DATA ANALYSIS (EDA)
# ============================================================================

def restaurant_detailed_analysis(df, total_original):
    """
    Detailed statistical analysis for each restaurant

    Parameters:
    - df: Cleaned DataFrame
    - total_original: Original order count for percentage calculations
    """
    print("\n" + "=" * 80)
    print("RESTAURANT DETAILED ANALYSIS")
    print("=" * 80)

    restaurant_stats = []

    for restaurant in sorted(df['Restaurant name'].unique()):
        rest_df = df[df['Restaurant name'] == restaurant]

        # Basic metrics
        total_orders = len(rest_df)
        unique_customers = rest_df['Customer ID'].nunique()
        market_share = (total_orders / len(df)) * 100

        # Order structure analysis
        multi_item_orders = 0
        single_item_orders = 0
        total_items = 0
        all_items = []

        for items in rest_df['Items in order'].dropna():
            item_list = str(items).split(', ')
            item_count = len(item_list)
            total_items += item_count

            if item_count > 1:
                multi_item_orders += 1
            else:
                single_item_orders += 1

            # Parse items with quantities
            for item in item_list:
                if ' x ' in item:
                    parts = item.strip().split(' x ')
                    try:
                        quantity = int(parts[0])
                        product = parts[1].strip()
                        all_items.extend([product] * quantity)
                    except:
                        all_items.append(item.strip())
                else:
                    all_items.append(item.strip())

        # Product variety
        item_counts = Counter(all_items)
        unique_items = len(item_counts)

        # Multi-item percentage
        multi_pct = (multi_item_orders / total_orders) * 100

        # Tier classification and min support recommendation
        if total_orders >= 5000 and multi_pct >= 50:
            rec_support = 0.01
            status = "Very Strong"
            tier = "Tier 1: Premium"
            rationale = f"High volume ({total_orders:,} orders) with strong multi-item rate ({multi_pct:.1f}%)"
        elif total_orders >= 1000 and multi_pct >= 40:
            rec_support = 0.01
            status = "Strong"
            tier = "Tier 1: Premium"
            rationale = f"Good volume ({total_orders:,} orders) with adequate multi-item rate ({multi_pct:.1f}%)"
        elif total_orders >= 200:
            rec_support = 0.05
            status = "Moderate"
            tier = "Tier 2: Limited"
            rationale = f"Limited volume ({total_orders:,} orders), higher threshold needed"
        elif total_orders >= 50:
            rec_support = None
            status = "Weak"
            tier = "Tier 3: Descriptive"
            rationale = f"Low volume ({total_orders:,} orders), only {multi_item_orders} multi-item orders"
        else:
            rec_support = None
            status = "Very Weak"
            tier = "Tier 3: Descriptive"
            rationale = f"Very low volume ({total_orders:,} orders), insufficient for association mining"

        restaurant_stats.append({
            'Restaurant': restaurant,
            'Tier': tier,
            'Total Orders': total_orders,
            'Pct of Dataset (%)': round((total_orders / len(df)) * 100, 2),
            'Market Share (%)': round(market_share, 2),
            'Multi-Item Orders': multi_item_orders,
            'Multi-Item (%)': round(multi_pct, 2),
            'Single-Item Orders': single_item_orders,
            'Single-Item (%)': round((single_item_orders / total_orders) * 100, 2),
            'Unique Products': unique_items,
            'Avg Items per Order': round(total_items / total_orders, 2),
            'Recommended min_support': rec_support,
            'Status': status,
            'Tier Rationale': rationale
        })

        print(f"\n{'=' * 80}")
        print(f"Restaurant: {restaurant}")
        print(f"{'=' * 80}")
        print(f"Tier: {tier}")
        print(f"Status: {status}")
        print(f"Rationale: {rationale}")
        print(f"\nOrder Statistics:")
        print(f"   Total Orders: {total_orders:,} ({(total_orders / len(df)) * 100:.2f}% of dataset)")
        print(f"   Multi-item orders: {multi_item_orders:,} ({multi_pct:.1f}%)")
        print(f"   Single-item orders: {single_item_orders:,} ({(single_item_orders / total_orders) * 100:.1f}%)")
        print(f"   Unique products: {unique_items}")
        print(f"   Avg items per order: {total_items / total_orders:.2f}")
        print(f"   Recommended min_support: {rec_support if rec_support else 'N/A (Descriptive only)'}")

        # Top 3 products
        if len(item_counts) > 0:
            print(f"\nTop 3 Products:")
            for i, (item, count) in enumerate(item_counts.most_common(3), 1):
                support = count / total_orders
                print(f"  {i}. {item}")
                print(f"     Sales: {count:,} | Support: {support:.4f} ({support * 100:.2f}%)")

    stats_df = pd.DataFrame(restaurant_stats)
    stats_df = stats_df.sort_values('Total Orders', ascending=False)

    print("\n" + "=" * 80)
    print("SUMMARY TABLE")
    print("=" * 80)
    print(stats_df[['Restaurant', 'Tier', 'Total Orders', 'Multi-Item (%)', 'Status']].to_string(index=False))

    return stats_df


def plot_eda_overview(stats_df):
    """
    Create EDA visualizations - FIXED LAYOUT
    Legend moved to bottom to prevent overlap.
    """
    fig = plt.figure(figsize=(20, 12))
    gs = fig.add_gridspec(2, 3, hspace=0.4, wspace=0.25)

    fig.suptitle('Restaurant Order Analysis - Exploratory Data Analysis',
                 fontsize=18, fontweight='bold', y=0.98)

    colors = ['#2ecc71', '#27ae60', '#f39c12', '#e67e22', '#e74c3c', '#c0392b']

    # 1. Order count
    ax1 = fig.add_subplot(gs[0, 0])
    bars = ax1.bar(range(len(stats_df)), stats_df['Total Orders'], color=colors, edgecolor='black', linewidth=1.5)
    ax1.set_ylabel('Number of Orders', fontweight='bold', fontsize=12)
    ax1.set_title('A) Order Volume by Restaurant', fontweight='bold', fontsize=13, pad=15)
    ax1.set_xticks(range(len(stats_df)))
    ax1.set_xticklabels(stats_df['Restaurant'], rotation=45, ha='right', fontsize=10)
    ax1.grid(axis='y', alpha=0.3)

    for bar in bars:
        ax1.text(bar.get_x() + bar.get_width() / 2., bar.get_height(),
                 f'{int(bar.get_height()):,}', ha='center', va='bottom', fontsize=9, fontweight='bold')

    # 2. Market share pie - FIXED LEGEND POSITION
    ax2 = fig.add_subplot(gs[0, 1])
    explode = [0.05 if i == 0 else 0 for i in range(len(stats_df))]

    wedges, texts, autotexts = ax2.pie(
        stats_df['Market Share (%)'],
        labels=None,  # Labels removed from pie to prevent overlap
        autopct=lambda pct: f'{pct:.1f}%' if pct > 5 else '',
        colors=colors,
        explode=explode,
        startangle=140,
        textprops={'fontsize': 11, 'weight': 'bold'},
        pctdistance=0.7
    )

    for autotext in autotexts:
        autotext.set_color('white')

    # Legend moved to BOTTOM
    legend_labels = [f"{row['Restaurant']}: {row['Market Share (%)']:.2f}%" for _, row in stats_df.iterrows()]
    ax2.legend(
        wedges, legend_labels, title="Restaurants",
        loc="upper center", bbox_to_anchor=(0.5, -0.05),  # Positioned below chart
        ncol=2, fontsize=9
    )
    ax2.set_title('B) Market Share Distribution', fontweight='bold', fontsize=13, pad=15)

    # 3. Multi-item order percentage
    ax3 = fig.add_subplot(gs[0, 2])
    bars = ax3.bar(range(len(stats_df)), stats_df['Multi-Item (%)'], color=colors, edgecolor='black', linewidth=1.5)
    ax3.axhline(y=50, color='red', linestyle='--', linewidth=2, label='Threshold: 50%')
    ax3.set_ylabel('Multi-Item Order Percentage (%)', fontweight='bold', fontsize=12)
    ax3.set_title('C) Multi-Item Order Percentage', fontweight='bold', fontsize=13, pad=15)
    ax3.set_xticks(range(len(stats_df)))
    ax3.set_xticklabels(stats_df['Restaurant'], rotation=45, ha='right', fontsize=10)
    ax3.legend(fontsize=10)
    ax3.grid(axis='y', alpha=0.3)
    ax3.set_ylim(0, max(stats_df['Multi-Item (%)']) * 1.2)

    for bar in bars:
        ax3.text(bar.get_x() + bar.get_width() / 2., bar.get_height(),
                 f'{bar.get_height():.1f}%', ha='center', va='bottom', fontsize=9, fontweight='bold')

    # 4. Tier classification
    ax4 = fig.add_subplot(gs[1, 0])
    tier_counts = stats_df['Tier'].value_counts().sort_index()
    tier_colors = {'Tier 1: Premium': '#27ae60', 'Tier 2: Limited': '#f39c12', 'Tier 3: Descriptive': '#e74c3c'}
    bars = ax4.bar(range(len(tier_counts)), tier_counts.values,
                   color=[tier_colors[t] for t in tier_counts.index], edgecolor='black', linewidth=1.5)
    ax4.set_ylabel('Restaurant Count', fontweight='bold', fontsize=12)
    ax4.set_title('D) Tier Distribution', fontweight='bold', fontsize=13, pad=15)
    ax4.set_xticks(range(len(tier_counts)))
    ax4.set_xticklabels(tier_counts.index, rotation=45, ha='right', fontsize=10)
    ax4.grid(axis='y', alpha=0.3)

    for bar in bars:
        ax4.text(bar.get_x() + bar.get_width() / 2., bar.get_height(),
                 f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=11, fontweight='bold')

    # 5. Unique product count
    ax5 = fig.add_subplot(gs[1, 1])
    bars = ax5.bar(range(len(stats_df)), stats_df['Unique Products'], color=colors, edgecolor='black', linewidth=1.5)
    ax5.set_ylabel('Unique Product Count', fontweight='bold', fontsize=12)
    ax5.set_title('E) Product Variety', fontweight='bold', fontsize=13, pad=15)
    ax5.set_xticks(range(len(stats_df)))
    ax5.set_xticklabels(stats_df['Restaurant'], rotation=45, ha='right', fontsize=10)
    ax5.grid(axis='y', alpha=0.3)

    for bar in bars:
        ax5.text(bar.get_x() + bar.get_width() / 2., bar.get_height(),
                 f'{int(bar.get_height())}', ha='center', va='bottom', fontsize=9, fontweight='bold')

    # 6. Average items per order
    ax6 = fig.add_subplot(gs[1, 2])
    bars = ax6.bar(range(len(stats_df)), stats_df['Avg Items per Order'], color=colors, edgecolor='black',
                   linewidth=1.5)
    ax6.set_ylabel('Average Items per Order', fontweight='bold', fontsize=12)
    ax6.set_title('F) Average Items per Order', fontweight='bold', fontsize=13, pad=15)
    ax6.set_xticks(range(len(stats_df)))
    ax6.set_xticklabels(stats_df['Restaurant'], rotation=45, ha='right', fontsize=10)
    ax6.grid(axis='y', alpha=0.3)

    for bar in bars:
        ax6.text(bar.get_x() + bar.get_width() / 2., bar.get_height(),
                 f'{bar.get_height():.2f}', ha='center', va='bottom', fontsize=9, fontweight='bold')

    plt.savefig('01_EDA_Overview.png', dpi=300, bbox_inches='tight')
    print("\nVisualization saved: 01_EDA_Overview.png")
    plt.close()


# ============================================================================
# PART 3: TRANSACTION FORMAT CONVERSION
# ============================================================================

def parse_items(items_string):
    """Parse 'Items in order' column"""
    items = []
    if pd.isna(items_string):
        return []

    for item in str(items_string).split(', '):
        item = item.strip()
        if ' x ' in item:
            parts = item.split(' x ')
            try:
                quantity = int(parts[0].strip())
                product = parts[1].strip()
                items.extend([product] * quantity)
            except:
                items.append(item)
        else:
            items.append(item)

    return items


def create_transactions(df, restaurant_name):
    """Create transaction list for a restaurant"""
    rest_df = df[df['Restaurant name'] == restaurant_name]
    transactions = []

    for items in rest_df['Items in order']:
        item_list = parse_items(items)
        if len(item_list) > 0:
            transactions.append(item_list)

    return transactions


def create_binary_matrix(transactions):
    """Convert transaction list to binary matrix"""
    te = TransactionEncoder()
    te_ary = te.fit(transactions).transform(transactions)
    df_encoded = pd.DataFrame(te_ary, columns=te.columns_)
    return df_encoded


# ============================================================================
# PART 4 & 5: APRIORI AND FP-GROWTH ALGORITHMS
# ============================================================================

def run_association_mining(df, restaurant_name, min_support=0.01,
                           min_confidence=0.20, min_lift=1.0, algorithm='Apriori'):
    """
    Run association rule mining with either Apriori or FP-Growth

    ADJUSTED PARAMETERS:
    - min_confidence lowered to 0.20 (from 0.25) for more rules
    - min_lift lowered to 1.0 (from 1.2) to capture more patterns
    """
    print(f"\n{'=' * 80}")
    print(f"{algorithm.upper()} ANALYSIS: {restaurant_name}")
    print(f"{'=' * 80}")
    print(f"Parameters: min_support={min_support}, min_confidence={min_confidence}, min_lift={min_lift}")

    transactions = create_transactions(df, restaurant_name)
    total_orders = len(df[df['Restaurant name'] == restaurant_name])

    print(f"\nDataset: {total_orders:,} total orders")
    print(f"Transactions for mining: {len(transactions):,}")
    print(f"Transaction rate: {(len(transactions) / total_orders) * 100:.1f}%")

    if len(transactions) == 0:
        print("No transactions found!")
        return None, 0, None

    df_encoded = create_binary_matrix(transactions)
    print(f"Unique items: {len(df_encoded.columns)}")

    print(f"\nRunning {algorithm} algorithm...")
    start_time = time.time()

    try:
        if algorithm == 'Apriori':
            frequent_itemsets = apriori(df_encoded, min_support=min_support, use_colnames=True)
        else:  # FP-Growth
            frequent_itemsets = fpgrowth(df_encoded, min_support=min_support, use_colnames=True)

        exec_time = time.time() - start_time

        print(f"Frequent itemsets found: {len(frequent_itemsets):,}")
        print(f"Execution time: {exec_time:.3f} seconds")

        if len(frequent_itemsets) > 0:
            rules = association_rules(
                frequent_itemsets,
                metric="confidence",
                min_threshold=min_confidence
            )

            # Apply lift filter
            rules = rules[rules['lift'] >= min_lift]
            rules = rules.sort_values('lift', ascending=False)

            print(f"Association rules generated: {len(rules):,}")
            print(f"Rules after filters (conf>={min_confidence}, lift>={min_lift}): {len(rules):,}")

            if len(rules) > 0:
                print(f"\nTop 5 Rules (sorted by lift):")
                print("-" * 80)

                for idx, row in rules.head(5).iterrows():
                    antecedents = ', '.join(list(row['antecedents']))
                    consequents = ', '.join(list(row['consequents']))
                    print(f"\nRule: {antecedents} => {consequents}")
                    print(f"  Support: {row['support']:.4f} ({row['support'] * 100:.2f}%)")
                    print(f"  Confidence: {row['confidence']:.4f} ({row['confidence'] * 100:.2f}%)")
                    print(f"  Lift: {row['lift']:.4f}")
            else:
                print("\nNo rules found meeting the criteria.")
                print(
                    f"Suggestion: Try lowering min_confidence (current: {min_confidence}) or min_lift (current: {min_lift})")

            return rules, exec_time, frequent_itemsets
        else:
            print("No frequent itemsets found!")
            print(f"Suggestion: Lower min_support (current: {min_support})")
            return None, exec_time, None

    except Exception as e:
        print(f"Error: {str(e)}")
        return None, 0, None


# ============================================================================
# PART 6: ALGORITHM COMPARISON
# ============================================================================

def compare_algorithms(df, total_orders):
    """Compare Apriori vs FP-Growth with detailed metrics"""
    print("\n\n" + "=" * 80)
    print("ALGORITHM COMPARISON: APRIORI vs FP-GROWTH")
    print("=" * 80)

    # Configuration for each tier
    restaurants_config = {
        'Aura Pizzas': {'min_support': 0.01, 'tier': 'Tier 1', 'min_confidence': 0.20, 'min_lift': 1.0},
        'Swaad': {'min_support': 0.008, 'tier': 'Tier 1', 'min_confidence': 0.20, 'min_lift': 1.0},
        # Lowered for more rules
        'Dilli Burger Adda': {'min_support': 0.05, 'tier': 'Tier 2', 'min_confidence': 0.20, 'min_lift': 1.0}
    }

    results = []

    for restaurant, config in restaurants_config.items():
        print(f"\n{'#' * 80}")
        print(f"COMPARING: {restaurant} ({config['tier']})")
        print(f"{'#' * 80}")

        rest_orders = len(df[df['Restaurant name'] == restaurant])

        # Apriori
        rules_apr, time_apr, freq_apr = run_association_mining(
            df, restaurant,
            min_support=config['min_support'],
            min_confidence=config['min_confidence'],
            min_lift=config['min_lift'],
            algorithm='Apriori'
        )

        # FP-Growth
        rules_fp, time_fp, freq_fp = run_association_mining(
            df, restaurant,
            min_support=config['min_support'],
            min_confidence=config['min_confidence'],
            min_lift=config['min_lift'],
            algorithm='FP-Growth'
        )

        # Calculate metrics
        speedup = time_apr / time_fp if time_fp > 0 else 0

        result = {
            'Restaurant': restaurant,
            'Tier': config['tier'],
            'Total Orders': rest_orders,
            'Orders Used (%)': 100.0,  # All delivered orders used
            'Min Support': config['min_support'],
            'Min Confidence': config['min_confidence'],
            'Min Lift': config['min_lift'],
            'Apriori Time (s)': round(time_apr, 3),
            'FP-Growth Time (s)': round(time_fp, 3),
            'Speedup (x)': round(speedup, 2),
            'Faster Algorithm': 'FP-Growth' if speedup > 1 else 'Apriori' if speedup < 1 else 'Equal',
            'Apriori Rules': len(rules_apr) if rules_apr is not None else 0,
            'FP-Growth Rules': len(rules_fp) if rules_fp is not None else 0,
            'Rules Match': 'Yes' if (len(rules_apr) if rules_apr is not None else 0) == (
                len(rules_fp) if rules_fp is not None else 0) else 'No',
            'Apriori Itemsets': len(freq_apr) if freq_apr is not None else 0,
            'FP-Growth Itemsets': len(freq_fp) if freq_fp is not None else 0
        }
        results.append(result)

        print(f"\n{'=' * 80}")
        print(f"COMPARISON SUMMARY: {restaurant}")
        print(f"{'=' * 80}")
        print(f"Dataset: {rest_orders:,} orders ({(rest_orders / total_orders) * 100:.2f}% of total)")
        print(f"Apriori Time: {time_apr:.3f}s | Rules: {result['Apriori Rules']}")
        print(f"FP-Growth Time: {time_fp:.3f}s | Rules: {result['FP-Growth Rules']}")
        print(f"Speedup: {speedup:.2f}x ({result['Faster Algorithm']} is faster)")
        print(f"Rule Count Match: {result['Rules Match']}")

    comparison_df = pd.DataFrame(results)

    print(f"\n\n{'=' * 80}")
    print("FINAL COMPARISON TABLE")
    print(f"{'=' * 80}")
    print(comparison_df[['Restaurant', 'Tier', 'Total Orders', 'Apriori Rules', 'FP-Growth Rules',
                         'Apriori Time (s)', 'FP-Growth Time (s)', 'Speedup (x)', 'Faster Algorithm']].to_string(
        index=False))

    # Save with metadata
    metadata = {
        'Total Dataset Orders': total_orders,
        'Analysis Date': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Algorithms Compared': 'Apriori vs FP-Growth',
        'Note': 'Lower parameters used to generate more rules for Swaad'
    }

    save_excel_with_formatting(comparison_df, '02_Algorithm_Comparison.xlsx',
                               sheet_name='Comparison', add_metadata=metadata)

    return comparison_df


def plot_algorithm_comparison(comparison_df):
    """Algorithm comparison visualizations"""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Apriori vs FP-Growth Algorithm Comparison',
                 fontsize=18, fontweight='bold')

    x = np.arange(len(comparison_df))
    width = 0.35

    # 1. Execution time
    ax1 = axes[0, 0]
    bars1 = ax1.bar(x - width / 2, comparison_df['Apriori Time (s)'], width,
                    label='Apriori', color='#3498db', edgecolor='black', linewidth=1.5)
    bars2 = ax1.bar(x + width / 2, comparison_df['FP-Growth Time (s)'], width,
                    label='FP-Growth', color='#e74c3c', edgecolor='black', linewidth=1.5)
    ax1.set_xlabel('Restaurant', fontweight='bold', fontsize=12)
    ax1.set_ylabel('Execution Time (seconds)', fontweight='bold', fontsize=12)
    ax1.set_title('A) Execution Time Comparison', fontweight='bold', fontsize=13, pad=15)
    ax1.set_xticks(x)
    ax1.set_xticklabels(comparison_df['Restaurant'], fontsize=10)
    ax1.legend(fontsize=11)
    ax1.grid(axis='y', alpha=0.3)

    for bars in [bars1, bars2]:
        for bar in bars:
            height = bar.get_height()
            ax1.text(bar.get_x() + bar.get_width() / 2., height,
                     f'{height:.3f}', ha='center', va='bottom', fontsize=8)

    # 2. Speedup
    ax2 = axes[0, 1]
    colors = ['#27ae60' if s > 1 else '#e74c3c' if s < 1 else '#95a5a6'
              for s in comparison_df['Speedup (x)']]
    bars = ax2.bar(comparison_df['Restaurant'], comparison_df['Speedup (x)'],
                   color=colors, edgecolor='black', linewidth=1.5)
    ax2.axhline(y=1, color='black', linestyle='--', linewidth=2, label='Equal Performance')
    ax2.set_xlabel('Restaurant', fontweight='bold', fontsize=12)
    ax2.set_ylabel('Speedup (Apriori Time / FP-Growth Time)', fontweight='bold', fontsize=12)
    ax2.set_title('B) Performance Speedup\n(Green: FP-Growth Faster | Red: Apriori Faster)',
                  fontweight='bold', fontsize=13, pad=15)
    ax2.legend(fontsize=11)
    ax2.grid(axis='y', alpha=0.3)

    for i, bar in enumerate(bars):
        height = bar.get_height()
        faster = comparison_df.iloc[i]['Faster Algorithm']
        speedup_val = comparison_df.iloc[i]['Speedup (x)']

        # Dynamically show which is faster
        if speedup_val > 1:
            label_text = f'{height:.2f}x\n(FP-Growth\nfaster)'
            label_color = '#27ae60'
        elif speedup_val < 1:
            actual_speedup = 1 / speedup_val  # Show how much faster Apriori is
            label_text = f'{height:.2f}x\n(Apriori\n{actual_speedup:.2f}x faster)'
            label_color = '#e74c3c'
        else:
            label_text = f'{height:.2f}x\n(Equal)'
            label_color = '#95a5a6'

        ax2.text(bar.get_x() + bar.get_width() / 2., height,
                 label_text, ha='center', va='bottom',
                 fontweight='bold', fontsize=8, color=label_color)

    # 3. Rules generated
    ax3 = axes[1, 0]
    bars1 = ax3.bar(x - width / 2, comparison_df['Apriori Rules'], width,
                    label='Apriori', color='#3498db', edgecolor='black', linewidth=1.5)
    bars2 = ax3.bar(x + width / 2, comparison_df['FP-Growth Rules'], width,
                    label='FP-Growth', color='#e74c3c', edgecolor='black', linewidth=1.5)
    ax3.set_xlabel('Restaurant', fontweight='bold', fontsize=12)
    ax3.set_ylabel('Number of Association Rules', fontweight='bold', fontsize=12)
    ax3.set_title('C) Association Rules Generated', fontweight='bold', fontsize=13, pad=15)
    ax3.set_xticks(x)
    ax3.set_xticklabels(comparison_df['Restaurant'], fontsize=10)
    ax3.legend(fontsize=11)
    ax3.grid(axis='y', alpha=0.3)

    for bars in [bars1, bars2]:
        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax3.text(bar.get_x() + bar.get_width() / 2., height,
                         f'{int(height)}', ha='center', va='bottom', fontsize=9)

    # 4. Frequent itemsets
    ax4 = axes[1, 1]
    bars1 = ax4.bar(x - width / 2, comparison_df['Apriori Itemsets'], width,
                    label='Apriori', color='#3498db', edgecolor='black', linewidth=1.5)
    bars2 = ax4.bar(x + width / 2, comparison_df['FP-Growth Itemsets'], width,
                    label='FP-Growth', color='#e74c3c', edgecolor='black', linewidth=1.5)
    ax4.set_xlabel('Restaurant', fontweight='bold', fontsize=12)
    ax4.set_ylabel('Number of Frequent Itemsets', fontweight='bold', fontsize=12)
    ax4.set_title('D) Frequent Itemsets Found', fontweight='bold', fontsize=13, pad=15)
    ax4.set_xticks(x)
    ax4.set_xticklabels(comparison_df['Restaurant'], fontsize=10)
    ax4.legend(fontsize=11)
    ax4.grid(axis='y', alpha=0.3)

    for bars in [bars1, bars2]:
        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax4.text(bar.get_x() + bar.get_width() / 2., height,
                         f'{int(height)}', ha='center', va='bottom', fontsize=8)

    plt.tight_layout()
    plt.savefig('03_Algorithm_Comparison_Charts.png', dpi=300, bbox_inches='tight')
    print("Visualization saved: 03_Algorithm_Comparison_Charts.png")
    plt.close()


# ============================================================================
# PART 7: BUSINESS INSIGHTS
# ============================================================================

def generate_combo_recommendations(rules, restaurant_name, total_orders, algorithm='Apriori', top_n=10):
    """Generate combo menu recommendations"""
    if rules is None or len(rules) == 0:
        print(f"No rules available for {restaurant_name} ({algorithm})")
        return None

    top_rules = rules.nlargest(min(top_n, len(rules)), 'lift')

    recommendations = []

    for i, (idx, row) in enumerate(top_rules.iterrows(), 1):
        antecedents = list(row['antecedents'])
        consequents = list(row['consequents'])
        combo_items = antecedents + consequents

        # Discount tiers
        if row['lift'] > 2.5:
            discount = "15-20%"
            priority = "HIGH"
        elif row['lift'] > 2.0:
            discount = "10-15%"
            priority = "MEDIUM"
        elif row['lift'] > 1.5:
            discount = "5-10%"
            priority = "LOW"
        else:
            discount = "5%"
            priority = "LOW"

        recommendations.append({
            'Restaurant': restaurant_name,
            'Algorithm': algorithm,
            'Rank': i,
            'Combo Items': ', '.join(combo_items),
            'Antecedents': ', '.join(antecedents),
            'Consequents': ', '.join(consequents),
            'Lift': round(row['lift'], 4),
            'Confidence (%)': round(row['confidence'] * 100, 2),
            'Support (%)': round(row['support'] * 100, 2),
            'Occurrences': int(row['support'] * total_orders),
            'Discount Recommendation': discount,
            'Priority': priority
        })

    return pd.DataFrame(recommendations)


# ============================================================================
# PART 8: MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function"""
    print("=" * 80)
    print("RESTAURANT ORDER ANALYSIS - ASSOCIATION RULE MINING PROJECT")
    print("Student: Toprak Kamburoğlu")
    print("=" * 80)

    # STEP 1: Data loading and cleaning
    print("\nSTEP 1: DATA LOADING AND CLEANING")
    print("-" * 80)

    df, metadata = load_and_clean_data('order_history_kaggle_data.csv')

    # Save cleaned data as Excel
    save_excel_with_formatting(df, 'cleaned_data_6_restaurants.xlsx',
                               sheet_name='Cleaned Data', add_metadata=metadata)
    # df.to_csv('cleaned_data_6_restaurants.csv', index=False)  # CSV disabled
    print("Cleaned data saved: cleaned_data_6_restaurants.xlsx")

    # STEP 2: EDA
    print("\nSTEP 2: EXPLORATORY DATA ANALYSIS")
    print("-" * 80)

    stats_df = restaurant_detailed_analysis(df, metadata['Original Orders'])

    eda_metadata = {
        'Total Orders in Analysis': len(df),
        'Original Dataset': metadata['Original Orders'],
        'Data Retention Rate': f"{metadata['Data Retention (%)']}%",
        'Number of Restaurants': stats_df['Restaurant'].nunique(),
        'Analysis Date': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    save_excel_with_formatting(stats_df, '00_Restaurant_Statistics.xlsx',
                               sheet_name='Statistics', add_metadata=eda_metadata)

    plot_eda_overview(stats_df)

    # STEP 3: Association mining & comparison
    print("\nSTEP 3: ASSOCIATION RULE MINING & ALGORITHM COMPARISON")
    print("-" * 80)

    comparison_df = compare_algorithms(df, len(df))
    plot_algorithm_comparison(comparison_df)

    # STEP 4: Business insights
    print("\nSTEP 4: BUSINESS INSIGHTS - COMBO RECOMMENDATIONS")
    print("-" * 80)

    all_recommendations = []

    # Tier 1 & 2 restaurants
    analysis_config = {
        'Aura Pizzas': {'min_support': 0.01, 'min_confidence': 0.20, 'min_lift': 1.0, 'top_n': 10},
        'Swaad': {'min_support': 0.008, 'min_confidence': 0.20, 'min_lift': 1.0, 'top_n': 10},
        'Dilli Burger Adda': {'min_support': 0.05, 'min_confidence': 0.20, 'min_lift': 1.0, 'top_n': 5}
    }

    for restaurant, config in analysis_config.items():
        rest_orders = len(df[df['Restaurant name'] == restaurant])

        print(f"\n{'#' * 80}")
        print(f"GENERATING RECOMMENDATIONS: {restaurant}")
        print(f"{'#' * 80}")

        # Apriori
        rules_apr, _, _ = run_association_mining(
            df, restaurant,
            min_support=config['min_support'],
            min_confidence=config['min_confidence'],
            min_lift=config['min_lift'],
            algorithm='Apriori'
        )

        if rules_apr is not None and len(rules_apr) > 0:
            rec = generate_combo_recommendations(rules_apr, restaurant, rest_orders, 'Apriori', config['top_n'])
            if rec is not None:
                all_recommendations.append(rec)

        # FP-Growth
        rules_fp, _, _ = run_association_mining(
            df, restaurant,
            min_support=config['min_support'],
            min_confidence=config['min_confidence'],
            min_lift=config['min_lift'],
            algorithm='FP-Growth'
        )

        if rules_fp is not None and len(rules_fp) > 0:
            rec = generate_combo_recommendations(rules_fp, restaurant, rest_orders, 'FP-Growth', config['top_n'])
            if rec is not None:
                all_recommendations.append(rec)

    # Save recommendations
    if len(all_recommendations) > 0:
        final_recommendations = pd.concat(all_recommendations, ignore_index=True)

        rec_metadata = {
            'Total Recommendations': len(final_recommendations),
            'Restaurants Analyzed': final_recommendations['Restaurant'].nunique(),
            'Algorithms Used': 'Apriori, FP-Growth',
            'Generated Date': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        save_excel_with_formatting(final_recommendations, '04_Combo_Menu_Recommendations.xlsx',
                                   sheet_name='Recommendations', add_metadata=rec_metadata)

    # STEP 5: Tier 3 analysis
    print("\nSTEP 5: TIER 3 RESTAURANTS - DESCRIPTIVE ANALYSIS")
    print("-" * 80)

    tier3_restaurants = ['Tandoori Junction', 'The Chicken Junction', 'Masala Junction']
    tier3_results = []

    for restaurant in tier3_restaurants:
        rest_df = df[df['Restaurant name'] == restaurant]

        print(f"\n{restaurant}: {len(rest_df)} orders (Descriptive only - insufficient for association mining)")

        all_items = []
        for items in rest_df['Items in order'].dropna():
            all_items.extend(parse_items(items))

        item_counts = Counter(all_items)

        for i, (item, count) in enumerate(item_counts.most_common(5), 1):
            support = count / len(rest_df)
            tier3_results.append({
                'Restaurant': restaurant,
                'Rank': i,
                'Product': item,
                'Sales Count': count,
                'Support': round(support, 4),
                'Support (%)': round(support * 100, 2),
                'Orders in Restaurant': len(rest_df)
            })

    tier3_df = pd.DataFrame(tier3_results)

    tier3_metadata = {
        'Note': 'These restaurants have insufficient data for association mining',
        'Reason': 'Low order volume and/or low multi-item order percentage',
        'Analysis Type': 'Frequency-based top products only',
        'Total Tier 3 Orders': sum(len(df[df['Restaurant name'] == r]) for r in tier3_restaurants)
    }

    save_excel_with_formatting(tier3_df, '05_Tier3_Top_Products.xlsx',
                               sheet_name='Top Products', add_metadata=tier3_metadata)

    # STEP 6: BUSINESS IMPACT ANALYSIS
    print(f"\n[DEBUG] Total recommendations collected: {len(all_recommendations)}")

    if len(all_recommendations) > 0:
        print(f"[DEBUG] Final recommendations shape: {final_recommendations.shape}")
        print(f"[DEBUG] Restaurants in recommendations: {final_recommendations['Restaurant'].unique()}")

        print("\n" + "=" * 80)
        print("STEP 6: BUSINESS IMPACT ANALYSIS")
        print("=" * 80)
        print("\nCalculating revenue potential, customer satisfaction, and ROI...")

        try:
            impact_df, summary_df = run_business_impact_analysis(
                df, final_recommendations
            )

            print("\nBusiness impact analysis completed!")
            print("  Generated files:")
            print("    - 06_Business_Impact_Analysis.xlsx")
            print("    - 07_Business_Impact_Visualization.png")
        except Exception as e:
            print(f"\n[ERROR] Business impact failed: {str(e)}")
            import traceback
            traceback.print_exc()
            print("Continuing without business impact...")
    else:
        print("\n[WARNING] No recommendations generated - Skipping business impact analysis")
        print("Possible reasons:")
        print("  1. No association rules found in any restaurant")
        print("  2. All rules filtered out by confidence/lift thresholds")
        print("  3. Algorithm errors occurred")

    # FINAL SUMMARY
    print("\n\n" + "=" * 80)
    print("ANALYSIS COMPLETED SUCCESSFULLY!")
    print("=" * 80)

    print("\nGenerated Files:")
    print("  1. cleaned_data_6_restaurants.xlsx - Cleaned dataset")
    print("  2. 00_Restaurant_Statistics.xlsx - Restaurant statistics")
    print("  3. 01_EDA_Overview.png - EDA visualizations")
    print("  4. 02_Algorithm_Comparison.xlsx - Comparison results")
    print("  5. 03_Algorithm_Comparison_Charts.png - Comparison charts")
    print("  6. 04_Combo_Menu_Recommendations.xlsx - Business recommendations")
    print("  7. 05_Tier3_Top_Products.xlsx - Tier 3 top products")
    if len(all_recommendations) > 0:
        print("  8. 06_Business_Impact_Analysis.xlsx - Business impact analysis")
        print("  9. 07_Business_Impact_Visualization.png - Business impact charts")

    print("\nKey Findings:")
    print(f"  Total Orders Analyzed: {len(df):,} (from {metadata['Original Orders']:,} original)")
    print(f"  Tier 1 Restaurants: 2 (Full association mining)")
    print(f"  Tier 2 Restaurants: 1 (Limited association mining)")
    print(f"  Tier 3 Restaurants: 3 (Descriptive only)")

    if len(comparison_df) > 0:
        avg_speedup = comparison_df['Speedup (x)'].mean()
        total_rules = comparison_df['Apriori Rules'].sum() + comparison_df['FP-Growth Rules'].sum()
        print(f"  Average Speedup: {avg_speedup:.2f}x")
        print(f"  Total Association Rules: {total_rules}")

    print("\nProject completed successfully!")


if __name__ == "__main__":
    main()



